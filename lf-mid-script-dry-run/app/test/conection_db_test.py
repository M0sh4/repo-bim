"""
    CONEXION A LA TABLA DE DATOS
"""
import json as j
import sys as s
import pandas as pd
from openpyxl import load_workbook

def get_extract_data():
    """
        Extraer data de EWP de la tabla 
    """
    data_as_dict_list = ""
    excel_file = "script_exc_req_dry_run/dry_run/data_db/dry-run-peticiones.xlsx"
    try:
        data = pd.read_excel(excel_file)
        if data.empty:
            print("El archivo Excel está vacío.")
            s.exit(1)
        else:
            for col in data.select_dtypes(include=['datetime64[ns]']).columns:
                data[col] = data[col].astype(str)
            data_as_dict_list = j.dumps(data.to_dict(orient="records"))
    except FileNotFoundError:
        print(f"Error DR: El archivo {excel_file} no se encuentra.")
        s.exit(1)
    except ConnectionError as e:
        print(f"Error DR: {e}")
        s.exit(1)
    return data_as_dict_list

def set_update_register_mob(number_db, codestatus_ewp,request_mob, response_mob, time_out_response, datetime_request): 
    excel_file = "script_exc_req_dry_run/dry_run/data_db/dry-run-peticiones.xlsx"
   
    try:
        book = load_workbook(excel_file)
        sheet_name = 'Hoja1'
        sheet = book[sheet_name]
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]  
        df = df[1:]
        df.reset_index(drop=True, inplace=True)

        print("Forma del DataFrame:", df.shape)
        print("Columnas del DataFrame:", df.columns.tolist())
        df['status_code_ewp'] = pd.to_numeric(df['status_code_ewp'], errors='coerce').fillna(0).astype(int)
        df['status_code_comviva'] = pd.to_numeric(df['status_code_comviva'], errors='coerce').fillna(0).astype(int)
        df['flg_status_registry'] = pd.to_numeric(df['flg_status_registry'], errors='coerce').fillna(0).astype(int)

        df['request_comviva_json'] = df['request_comviva_json'].astype(object)
        df['response_comviva_json'] = df['response_comviva_json'].astype(object)
        df['date_test_execution'] = df['date_test_execution'].astype(object)
        df['is_result_equals'] = df['is_result_equals'].astype(object)
        df['time_execution'] = df['time_execution'].astype(object)

        status_code_ewp_col = df.columns.get_loc('status_code_ewp') + 1  
        flg_status_registry_col = df.columns.get_loc('flg_status_registry') + 1  
        request_comviva_col = df.columns.get_loc('request_comviva_json') + 1  
        response_comviva_col = df.columns.get_loc('response_comviva_json') + 1
        statuscode_comviva_col = df.columns.get_loc('status_code_comviva') + 1
        date_col = df.columns.get_loc('date_test_execution') + 1
        result_col = df.columns.get_loc('is_result_equals') + 1
        time_col = df.columns.get_loc('time_execution') + 1

        statuscode_comviva = int(response_mob["statusCode"])
        for index in df.index:
            try:
                if df.at[index, 'number'] == number_db:  
                    sheet.cell(row=index + 2, column=status_code_ewp_col, value=int(codestatus_ewp))
                    sheet.cell(row=index + 2, column=flg_status_registry_col, value=int(1))
                    sheet.cell(row=index + 2, column=request_comviva_col, value=str(j.dumps(request_mob))) 
                    sheet.cell(row=index + 2, column=response_comviva_col, value=j.dumps(response_mob)) 
                    sheet.cell(row=index + 2, column=statuscode_comviva_col, value=int(statuscode_comviva)) 
                    sheet.cell(row=index + 2, column=date_col, value=datetime_request)  
                    sheet.cell(row=index + 2, column=result_col, value=1 if True else 0)  
                    sheet.cell(row=index + 2, column=time_col, value=time_out_response)
            except Exception as e:
                print(f"Error con el índice {index}: {e}")
        book.save(excel_file)

    except PermissionError as pe:
        print(f"Error de permiso: {pe}")
    except Exception as e:
        print(f"Error inesperado: {e}")
get_extract_data()
